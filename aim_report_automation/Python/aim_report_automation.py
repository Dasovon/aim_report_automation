"""AIM Report Automation core workflow.

This module reads AiM browse/fc_review CSV exports and produces a
formatted Excel workbook tailored for the ETB/WEB/HEB Facilities
Coordinator workflow. Key capabilities include:

- Required fields (Building, Floor, Room, Age (Work Days), Inspection Status)
  with validation.
- Business-day aging using Date Created/Edit Date values.
- Robust floor/room extraction from Description text.
- Building/floor/room sorting rules specific to ETB, WEB, and HEB.
- Excel output with inspection dropdowns, conditional formatting, and
  touch-friendly layout.

Run as a script:

    python aim_report_automation.py input.csv output.xlsx

If no output path is provided a timestamped file is created alongside the
input CSV.
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BUILDING_ORDER = ["ETB", "WEB", "HEB", ""]
INSPECTION_STATUSES = ["Pending", "Complete", "Incomplete", "Needs Review"]


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------
def _clean_headers(columns: Iterable[str]) -> list[str]:
    """Normalize CSV headers by stripping whitespace and control characters."""

    cleaned = []
    for col in columns:
        cleaned.append(re.sub(r"[^\x20-\x7E]", "", str(col)).strip().replace("\ufeff", ""))
    return cleaned


def _find_column(df: pd.DataFrame, *search_terms: str) -> Optional[str]:
    """Return the first column whose name contains any of the search terms."""

    lower_cols = {c.lower(): c for c in df.columns}
    for term in search_terms:
        for col in df.columns:
            if term.lower() in col.lower():
                return col
        if term.lower() in lower_cols:
            return lower_cols[term.lower()]
    return None


def _normalize_building(df: pd.DataFrame) -> pd.Series:
    """Derive the Building column from the most relevant source."""

    building_col = _find_column(df, "building", "property", "facility")
    if building_col is None:
        return pd.Series(["" for _ in range(len(df))], index=df.index)

    series = df[building_col].fillna("").astype(str).str.upper().str.strip()
    return series.replace({"E.T.B": "ETB", "W.E.B": "WEB", "H.E.B": "HEB"})


def _extract_floor_room(description: str) -> Tuple[str, str]:
    """Extract floor and room information from a description string."""

    if not isinstance(description, str) or not description.strip():
        return "", ""

    desc = description.strip()

    # Floor patterns
    floor_patterns = [
        r"(?:floor|flr|level|lvl)\s*[:\-]?\s*([A-Za-z0-9]+)",
        r"\b(LL|SF)\b",
    ]

    floor_val = ""
    for pattern in floor_patterns:
        match = re.search(pattern, desc, flags=re.IGNORECASE)
        if match:
            floor_val = match.group(1)
            break

    # Room patterns
    room_patterns = [
        r"(?:room|rm)\s*[:\-]?\s*([A-Za-z0-9]+)",
        r"\b(\d{3,4}[A-Z]?)\b",
    ]

    room_val = ""
    for pattern in room_patterns:
        match = re.search(pattern, desc, flags=re.IGNORECASE)
        if match:
            room_val = match.group(1)
            break

    # Fallback: derive floor from room leading digit(s)
    if not floor_val and room_val:
        if len(room_val) >= 2 and room_val[:2].isdigit():
            floor_val = room_val[:2] if room_val[:2] in {"10", "11", "12"} else room_val[0]
        elif room_val and room_val[0].isdigit():
            floor_val = room_val[0]

    if floor_val.upper() in {"0", "B"}:
        floor_val = "B"
    elif floor_val.upper() == "LL":
        floor_val = "LL"
    elif floor_val.upper() == "SF":
        floor_val = "SF"

    return floor_val, room_val


def _floor_rank(value: str) -> Tuple[int, str]:
    """Return a tuple used for floor sorting (blank -> B/LL/SF -> numeric)."""

    if value is None or str(value).strip() == "":
        return (0, "")

    val = str(value).upper().strip()
    if val in {"B", "LL", "SF"}:
        return (1, val)

    if val.isdigit():
        return (2, int(val))

    return (3, val)


def _room_rank(value: str) -> Tuple[int, str, str]:
    """Return a tuple used for room sorting (blank -> numeric -> alphanumeric)."""

    if value is None or str(value).strip() == "":
        return (0, "", "")

    val = str(value).upper().strip()
    numeric_match = re.match(r"(\d+)([A-Z]*)", val)
    if numeric_match:
        number_part = int(numeric_match.group(1))
        suffix = numeric_match.group(2)
        return (1, number_part, suffix)

    return (2, val, "")


def _building_rank(value: str) -> int:
    """Custom building rank: ETB -> WEB -> HEB -> blank -> others."""

    val = str(value).upper().strip() if value else ""
    if val in BUILDING_ORDER:
        return BUILDING_ORDER.index(val)
    if val == "":
        return len(BUILDING_ORDER)
    return len(BUILDING_ORDER) + 1


def _business_days(start: pd.Timestamp, end: pd.Timestamp) -> Optional[int]:
    """Calculate business-day difference between two timestamps."""

    if pd.isna(start):
        return None

    try:
        return int(np.busday_count(start.date(), end.date()))
    except Exception:
        return None


def _normalize_dates(df: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    """Normalize date columns to YYYY-MM-DD strings where possible."""

    for col in columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m-%d")
    return df


# ---------------------------------------------------------------------------
# Core data pipeline
# ---------------------------------------------------------------------------
def prepare_dataframe(csv_path: str) -> pd.DataFrame:
    """Load and transform the CSV into the standardized report dataframe."""

    df = pd.read_csv(csv_path)
    df.columns = _clean_headers(df.columns)

    # Identify relevant columns
    desc_col = _find_column(df, "description", "desc")
    date_created_col = _find_column(df, "date created", "created")
    edit_date_col = _find_column(df, "edit date", "last updated", "modified")

    if desc_col is None:
        raise ValueError("Missing required Description column in CSV input.")

    # Add derived fields
    df["Building"] = _normalize_building(df)
    df["Floor"], df["Room"] = zip(*df[desc_col].apply(_extract_floor_room))

    today = pd.Timestamp(datetime.now(timezone.utc).date())
    date_for_age = None
    if date_created_col:
        date_for_age = pd.to_datetime(df[date_created_col], errors="coerce")
    elif edit_date_col:
        date_for_age = pd.to_datetime(df[edit_date_col], errors="coerce")

    df["Age (Work Days)"] = (
        date_for_age.apply(lambda d: _business_days(d, today)) if date_for_age is not None else None
    )

    df["Inspection Status"] = "Pending"

    # Normalize date columns for readability
    df = _normalize_dates(df, filter(None, [date_created_col, edit_date_col]))

    # Filter to non-empty original columns so we avoid blank-only fields
    non_empty_cols = [c for c in df.columns if df[c].notna().any()]

    # Ensure required columns exist
    required = {"Building", "Floor", "Room", "Age (Work Days)", "Inspection Status"}
    missing_required = [c for c in required if c not in df.columns]
    if missing_required:
        raise ValueError(f"Missing required columns after processing: {missing_required}")

    # Sort records according to custom rules
    df["__building_rank"] = df["Building"].apply(_building_rank)
    df["__floor_rank"] = df["Floor"].apply(_floor_rank)
    df["__room_rank"] = df["Room"].apply(_room_rank)

    df.sort_values(by=["__building_rank", "__floor_rank", "__room_rank"], inplace=True)
    df.drop(columns=["__building_rank", "__floor_rank", "__room_rank"], inplace=True)

    # Reorder columns: required first, then remaining original columns
    ordered_cols = [c for c in ["Building", "Floor", "Room", "Age (Work Days)", "Inspection Status"] if c in df.columns]
    ordered_cols += [c for c in non_empty_cols if c not in ordered_cols]
    df = df[ordered_cols]

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def _apply_excel_formatting(df: pd.DataFrame, output_path: str) -> None:
    """Write the dataframe to Excel with formatting, validation, and CF."""

    output_path = str(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"

    # Write header and data
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row.values))

    # Freeze panes
    ws.freeze_panes = "B2"

    # Styles
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

        column_letter = get_column_letter(col_idx)
        if "description" in header.lower():
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = 18

    # Body formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            header_val = ws.cell(row=1, column=cell.column).value
            if header_val and "description" in str(header_val).lower():
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Data validation dropdown for Inspection Status
    if "Inspection Status" in df.columns:
        insp_col_idx = df.columns.get_loc("Inspection Status") + 1
        dv = DataValidation(type="list", formula1='"' + ",".join(INSPECTION_STATUSES) + '"', allow_blank=True)
        dv.error = "Please choose a valid inspection status."
        dv.prompt = "Select an inspection status."
        ws.add_data_validation(dv)
        dv_range = f"{ws.cell(row=2, column=insp_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=insp_col_idx).coordinate}"
        dv.add(dv_range)

        # Stoplight conditional formatting
        status_col_letter = get_column_letter(insp_col_idx)
        for status, color in [
            ("Complete", "C6EFCE"),
            ("Pending", "FFF2CC"),
            ("Incomplete", "F8CBAD"),
            ("Needs Review", "FFD966"),
        ]:
            formula = f'${status_col_letter}$2:${status_col_letter}${ws.max_row}'
            ws.conditional_formatting.add(
                formula,
                FormulaRule(formula=[f'${status_col_letter}2="{status}"'], stopIfTrue=False, fill=PatternFill(start_color=color, end_color=color, fill_type="solid")),
            )

    # Heatmap for Age (Work Days)
    if "Age (Work Days)" in df.columns:
        age_col_idx = df.columns.get_loc("Age (Work Days)") + 1
        age_range = f"{ws.cell(row=2, column=age_col_idx).coordinate}:{ws.cell(row=ws.max_row, column=age_col_idx).coordinate}"
        ws.conditional_formatting.add(
            age_range,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="C6EFCE",
                mid_type="num",
                mid_value=max(df["Age (Work Days)"].dropna().median(), 1) if df["Age (Work Days)"].notna().any() else 1,
                mid_color="FFF2CC",
                end_type="max",
                end_color="F4B084",
            ),
        )

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Command-line interface
# ---------------------------------------------------------------------------
def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Format AIM browse/fc_review CSV exports for ETB/WEB/HEB reporting.")
    parser.add_argument("csv_path", help="Path to the browse.csv or fc_review.csv file")
    parser.add_argument("output", nargs="?", help="Optional output Excel path (.xlsx or .xlsm)")
    return parser


def main(argv: Optional[list[str]] = None) -> None:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    csv_path = Path(args.csv_path).expanduser().resolve()
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    df = prepare_dataframe(str(csv_path))

    output = args.output
    if output:
        output_path = Path(output).expanduser().resolve()
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = csv_path.with_name(f"{timestamp}_aim_report.xlsx")

    _apply_excel_formatting(df, str(output_path))
    print(f"Saved formatted workbook to {output_path}")


if __name__ == "__main__":
    main()

