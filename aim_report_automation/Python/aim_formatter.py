import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from tkinter import Tk, filedialog
import re, os, subprocess, platform

# ------------------------------------------------------------
# Floor / Room Extraction
# ------------------------------------------------------------
def extract_floor_room(description: str):
    if not description or not isinstance(description, str):
        return "", ""
    desc = description.strip()
    floor_val, room_val = "", ""

    floor_match = re.search(r"(?:Floor:|Flr:)\s*([A-Za-z0-9]+)", desc, re.IGNORECASE)
    if floor_match:
        floor_val = floor_match.group(1).strip()

    room_match = re.search(r"(?:Room:|Rm:)\s*([A-Za-z0-9]+)", desc, re.IGNORECASE)
    if room_match:
        room_val = room_match.group(1).strip()

    # fallback from room number
    if not floor_val and room_val:
        if len(room_val) >= 4 and room_val[:2].isdigit() and room_val[:2] in ("10", "11", "12"):
            floor_val = room_val[:2]
        else:
            mapping = {"0": "B", "1": "1", "2": "2", "3": "3", "4": "4",
                       "5": "5", "6": "6", "7": "7", "8": "8", "9": "9"}
            floor_val = mapping.get(room_val[0], "")
    if floor_val.upper() == "SF":
        floor_val = "SF"
    elif floor_val in ("0", "B"):
        floor_val = "B"
    return floor_val, room_val


def floor_rank(floor_val):
    if not floor_val or str(floor_val).strip() == "":
        return 999
    v = str(floor_val).upper().strip()
    if v == "B":
        return 0
    if v == "SF":
        return 99
    return int(v) if v.isdigit() else 999


def room_rank(room_val):
    if not room_val or str(room_val).strip() == "":
        return 999999
    val = room_val.upper().strip()
    if any(x in val for x in ("HALL", "STR", "ELEV")):
        return 700000
    match = re.match(r"(\d+)", val)
    return int(match.group(1)) if match else 600000


def calculate_business_days(start_date, end_date):
    if pd.isna(start_date):
        return None
    try:
        return np.busday_count(start_date.date(), end_date.date()) or 0
    except Exception:
        return None


# ------------------------------------------------------------
# MAIN FORMATTER
# ------------------------------------------------------------
def run_aim_formatter(csv_path, output_path=None):
    print(f"\n=== Running AIM Formatter ===")
    print(f"Input CSV: {csv_path}\n")

    df = pd.read_csv(csv_path)
    today = datetime.now()

    # clean headers
    df.columns = [re.sub(r"[^\x20-\x7E]", "", c).strip().replace("\ufeff", "") for c in df.columns]
    print("Detected headers:", list(df.columns))

    def find_col(part):
        for c in df.columns:
            if part.lower() in c.lower():
                return c
        return None

    work_col = find_col("work")
    desc_col = find_col("desc")
    date_col = find_col("date")
    if not all([work_col, desc_col, date_col]):
        raise ValueError(f"Missing required columns. Found: {list(df.columns)}")

    df.rename(columns={work_col: "Work Order", desc_col: "Description",
                       date_col: "Date Created"}, inplace=True)

    # derived columns
    df["Age (Days)"] = df["Date Created"].apply(lambda d: calculate_business_days(pd.to_datetime(d), today))
    df["Floor"], df["Room"] = zip(*df["Description"].apply(extract_floor_room))
    df["__FloorRank"] = df["Floor"].apply(floor_rank)
    df["__RoomRank"] = df["Room"].apply(room_rank)
    df["Inspection Status"] = "Pending"
    df.sort_values(by=["__FloorRank", "__RoomRank"], inplace=True)

    # load template (using relative path to this script's location)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, "template.xlsm")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Work Orders"]
    ws.delete_rows(1, ws.max_row)

    headers = list(df.columns)
    ws.append(headers)

    # header style
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.value = h
        c.fill = header_fill
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # write data
    print(f"Writing {len(df)} rows to Excel...")
    for _, row in df.iterrows():
        ws.append(list(row.values))

    # remove helper cols
    room_idx = headers.index("__RoomRank") + 1
    floor_idx = headers.index("__FloorRank") + 1
    ws.delete_cols(room_idx)
    ws.delete_cols(floor_idx)

    # formatting
    ws.freeze_panes = "A2"
    for col in ws.columns:
        header = ws.cell(row=1, column=col[0].column)
        name = str(header.value)
        if "Description" in name:
            ws.column_dimensions[col[0].column_letter].width = 45
        elif "Inspection Status" in name:
            ws.column_dimensions[col[0].column_letter].width = 20
        else:
            ws.column_dimensions[col[0].column_letter].width = 14

    # dropdown
    insp_col = headers.index("Inspection Status") + 1
    dv = DataValidation(type="list",
                        formula1='"Pending,Complete,Incomplete,Needs Review"',
                        allow_blank=True)
    dv.error = "Please choose a valid status."
    dv.prompt = "Select an inspection status."
    ws.add_data_validation(dv)
    dv_range = f"{ws.cell(row=2, column=insp_col).coordinate}:{ws.cell(row=ws.max_row, column=insp_col).coordinate}"
    dv.add(dv_range)

    # borders & align
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            header_val = ws.cell(row=1, column=cell.column).value
            if header_val and "Description" in str(header_val):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # dashboard
    create_dashboard(ws, wb)

    # save
    if not output_path:
        downloads = os.path.expanduser("~/Downloads")
        today_str = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(downloads, f"{today_str}_WOs.xlsm")

    wb.save(output_path)
    print(f"✅ Saved macro-enabled workbook: {output_path}\n")

    if platform.system() == "Darwin":
        subprocess.run(["open", output_path])
    elif platform.system() == "Windows":
        os.startfile(output_path)


# ------------------------------------------------------------
# Dashboard sheet
# ------------------------------------------------------------
def create_dashboard(ws_main, wb):
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws_dash = wb.create_sheet("Dashboard")

    statuses = ["Pending", "Complete", "Incomplete", "Needs Review"]
    counts = {s: 0 for s in statuses}
    total_age = 0
    age_count = 0

    insp_col, age_col = None, None
    for i, header in enumerate(ws_main[1], start=1):
        if header.value == "Inspection Status":
            insp_col = i
        elif header.value == "Age (Days)":
            age_col = i
    if not insp_col or not age_col:
        return

    for row in ws_main.iter_rows(min_row=2, values_only=True):
        status = row[insp_col - 1]
        age = row[age_col - 1]
        if status in counts:
            counts[status] += 1
        if isinstance(age, (int, float)):
            total_age += age
            age_count += 1

    avg_age = round(total_age / age_count, 2) if age_count else 0
    ws_dash.append(["Status", "Count"])
    for s in statuses:
        ws_dash.append([s, counts[s]])
    ws_dash.append(["Average Age (Days)", avg_age])

    for col in ws_dash.columns:
        ws_dash.column_dimensions[col[0].column_letter].width = 25
    for cell in ws_dash["A"] + ws_dash["B"]:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
    ws_dash["A1"].font = Font(bold=True)
    ws_dash["B1"].font = Font(bold=True)


# ------------------------------------------------------------
# Entry point
# ------------------------------------------------------------
if __name__ == "__main__":
    Tk().withdraw()
    downloads = os.path.expanduser("~/Downloads")
    today_str = datetime.now().strftime("%Y%m%d")
    default_filename = f"{today_str}_WOs.xlsm"

    file_path = filedialog.askopenfilename(
        title="Select browse.csv file",
        initialdir=downloads,
        filetypes=[("CSV files", "*.csv")]
    )

    if file_path:
        save_path = filedialog.asksaveasfilename(
            title="Save formatted file as",
            initialdir=downloads,
            initialfile=default_filename,
            defaultextension=".xlsm",
            filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm")]
        )
        run_aim_formatter(file_path, save_path)
